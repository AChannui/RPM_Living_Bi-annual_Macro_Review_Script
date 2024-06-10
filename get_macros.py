import argparse
import datetime
import os

from collections import defaultdict
from datetime import timedelta
from pprint import pprint

import requests_cache

from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill

# no logging just print debugging

def main():
    # argument parsing
    parser = argparse.ArgumentParser()
    parser.add_argument("--user", "-u", type=str, action="store", default=os.getenv("ZENDESK_USERNAME"))
    parser.add_argument("--password", "-p", type=str, action="store", default=os.getenv("ZENDESK_PASSWORD"))
    args = parser.parse_args()

    # request macros from zendesk
    # session = requests.Session()
    session = requests_cache.CachedSession('cache')
    session.auth = (args.user, args.password)
    # print debugging
    # print(session.auth)
    next_url = "https://roscoeproperties.zendesk.com/api/v2/macros.json?per_page=200&include=usage_30d"
    groups_url = "https://roscoeproperties.zendesk.com/api/v2/groups"
    
    # getting all macros and narrowing down to all active macros
    macros = get_macro_list(session, next_url)
    active_macros = [item for item in macros if item["active"]]
    # print debugging
    # print(len(active_macros))

    # getting all group ids
    groups = get_macro_list(session, groups_url, "groups")
    # print debugging
    # print(len(groups))
    # dict of group id and corresponding name
    public_macro_sheet_name = "Public Shared Macros"
    group_map = {item["id"]:item["name"].replace("/", "") for item in groups if not item["deleted"]}
    group_map.update({-1:public_macro_sheet_name})

    # creates dict of macros by group id
    grouped_macros = defaultdict(list)
    sort_macros(active_macros, grouped_macros)


    # creates workbook with each group having their own sheet
    wb = Workbook()
    create_workbook(group_map, grouped_macros, wb)
        
    #print debugging
    # print(wb.get_sheet_names())
    del wb['Sheet']
    # print debugging
    # print(wb.get_sheet_names())

    wb._sheets.sort(key=lambda ws: ws.title)
    move_public_macro_sheet(wb, public_macro_sheet_name)
    wb.save("macro_review.xlsx")

    # print debugging
    # print(len(macros))
    print('done')

# moves sheet with matching name to front.
def move_public_macro_sheet(wb, public_macro_sheet_name):
    for count, sheet in enumerate(wb._sheets):
        # print debugging
        # print(count, sheet.title)
        if sheet.title == public_macro_sheet_name:
            temp = wb._sheets.pop(count)
            wb._sheets.insert(0, temp)
    
                

def sort_macros(active_macros, grouped_macros):
    for macro in active_macros:
        restriction = macro["restriction"]
        # save off none restrictions to shared macros group
        if restriction is None:
            macro["restriction"] = -1
            print(f"restriction is null, macro id number: {macro['id']}")
            grouped_macros[-1].append(macro)
            continue

        # filter out non group restrictions
        if restriction["type"] != "Group":
            print(f"restriction not group, macro id number: {macro['id']}")
            continue
        # creates new 
        for id in restriction['ids']:
            grouped_macros[id].append(macro)
            

def create_workbook(group_map, grouped_macros, wb):
    for id, macro_list in grouped_macros.items():
        # sheet creation
        # print debugging
        # print(group_map[id])
        ws1 = wb.create_sheet(group_map[id])
        # adding header
        header = ['Name', "ID", "Created", "Updated", "Group", "Usage 30 Days", "Action", "Action Taken"]
        ws1.append(header)

        #set font
        font = Font(bold=True, underline='single')
        for cell in ws1[1]:
            cell.font = font
            
        # highlight set up
        # no action green 0099CC00 50
        # review yellow 00FFFF99 43
        # deactivate orange 00FFCC99 52
        no_action_index = Color(indexed=50)
        review_index = Color(indexed=43)
        deactivation_index = Color(indexed=52)
        highlight_review = PatternFill(patternType='solid', fgColor=review_index)
        highlight_deactivation = PatternFill(patternType='solid', fgColor=deactivation_index)
        highlight_no_action = PatternFill(patternType='solid', fgColor=no_action_index)

        # date operations set up
        comparision_date = datetime.datetime.today() - timedelta(days=90)
        # print debugging
        # print(comparision_date)

        # populating sheet with associated macros and highlighting macros of interest
        for macro in macro_list:
            # setting none restrictions to n/a group name
            if macro['restriction'] == -1:
                macro_groups = "N/A"
            else:
                macro_groups = ", ".join (group_map[item] for item in macro['restriction']['ids'])
            updated_time = datetime.datetime.strptime(macro['updated_at'], '%Y-%m-%dT%H:%M:%SZ')

            if(updated_time < comparision_date and macro['usage_30d'] == 0):
                ws1.append([macro["title"], macro["id"], macro["created_at"], macro["updated_at"], macro_groups, macro["usage_30d"], "Deactivation Suggested"])
                for cell in ws1[ws1.max_row]:
                    cell.fill = highlight_deactivation
            
            elif(updated_time < comparision_date and macro['usage_30d'] > 0):
                ws1.append([macro["title"], macro["id"], macro["created_at"], macro["updated_at"], macro_groups, macro["usage_30d"], "Department Review"])
                for cell in ws1[ws1.max_row]:
                    cell.fill = highlight_review

            # don't think is necessary
            # elif(updated_time >= comparision_date and macro['usage_30d'] > 0):
            #     ws1.append([macro["title"], macro["id"], macro["created_at"], macro["updated_at"], macro_groups, macro["usage_30d"], "Department Review"])

            else:
                ws1.append([macro["title"], macro["id"], macro["created_at"], macro["updated_at"], macro_groups, macro["usage_30d"], "No Action, Recommend Review"])
                for cell in ws1[ws1.max_row]:
                    cell.fill = highlight_no_action

            # ws1.append([macro["title"], macro["id"], macro["created_at"], macro["updated_at"], macro_groups, macro["usage_30d"]])

            # print debugging
            # print(updated_time)
            # if macro['usage_30d'] is 0 or (updated_time < comparision_date and macro['usage_30d'] < 15):
            #     for cell in ws1[ws1.max_row]:
            #         cell.fill = highlight_fill

        # set id to show in regular form not scientific notation
        colB = ws1['B']
        for cell in colB:
            cell.number_format = '0'
        #hidding macro id becuase not shown included in past reviews but might be useful later on
        ws1.column_dimensions['B'].hidden=True

        # changing dates from iso dates to 3 letter months with number year and day
        colC = ws1['C']
        for cell in colC[1:]:
            cell.value = convert_iso_to_date(cell.value)

        colD = ws1['D']
        for cell in colD[1:]:
            cell.value = convert_iso_to_date(cell.value)
            
        auto_space_column_width(ws1)
        ws1.auto_filter.ref = ws1.dimensions    


def auto_space_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        worksheet.column_dimensions[column_letter].width = adjusted_width

# get cell value with ISOdate and converts it into a date
def convert_iso_to_date(cell_value):
    full_date = datetime.datetime.fromisoformat(cell_value)
    # print(dt)
    number_date = full_date.date()
    month_abbreviation = number_date.strftime('%Y-%b-%d')
    #print(month_abbreviation)
    return month_abbreviation

# grabs all data from zendesk from url with key to specify what to grab
def get_macro_list(session, next_url, key="macros"):
    results = []
    while next_url:
        pprint(next_url)
        response = session.get(next_url)
        response.raise_for_status()
        data = response.json()
        # print debugging
        # print(json.dumps(response.json(), indent=2))
        next_url = data['next_page']
        results.extend(data[key])
    return results


if __name__ == "__main__":
    main()
